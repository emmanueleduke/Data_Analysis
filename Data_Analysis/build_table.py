import openpyxl

workbook = openpyxl.load_workbook("excel_file_path")
sheet = workbook['Code with Simplearn']

# Update cell value
sheet.cell(row=1, column=1, value='Subject')
sheet.cell(row=2, column=1, value='Maths')
sheet.cell(row=3, column=1, value='English')
sheet.cell(row=4, column=1, value='Automation')
sheet.cell(row=5, column=1, value='Python')
sheet.cell(row=6, column=1, value='HTML')


# Save changes
workbook.save("excel_file_path")

# Close the workbook
workbook.close()

